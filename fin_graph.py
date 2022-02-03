from matplotlib import pyplot as plt
import fin_func
from openpyxl import load_workbook

# read Excel
wb = load_workbook(filename="fin.xlsx")

# sheet??
# print(wb.sheetnames)

# data?? 코스피, 코스닥, 다우존스, S&P 500, 나스닥	
kospi = wb['data']['A':'E']         # 코스피 [0]
kosdaq = wb['data']['H':'L']        # 코스닥 [1]
djia = wb['data']['O':'S']          # 나스닥 [2]
sp = wb['data']['V':'Z']            # 다우지수 [3]
nasdaq = wb['data']['AC':'AG']      # 스탠다드 푸어스 [4]
ku = wb['data']['AL':'AM']          # 환율(Korea - USA) [5]

# 시장
stocks = []
stocks.append(kospi)
stocks.append(kosdaq)
stocks.append(djia)
stocks.append(sp)
stocks.append(nasdaq)

# 데이터 가공
organStock = []
organExchange = []
for stock in stocks:
    # 코스피 코스닥 나스닥 다우지수 스탠다드
    # Date[0] Open[1] High[2] Low[3] Close[4]
    organStock.append(fin_func.stockDataOrgan(stock))
organExchange.append(fin_func.exchangeDataOrgan(ku))

# timeline setting
timeSetData = fin_func.timelineSet(organStock, organExchange)

# dimension
# organ[stock(market)][stock column][data]

timeline = organStock[3][0]
kospiClose = timeSetData[0]
kosdaqClose = timeSetData[1]
djiaClose = timeSetData[2]
spClose = timeSetData[3]
nasdaqClose = timeSetData[4]
kuClose = timeSetData[5]

# 그래프
fin_func.figureSet(timeline, [kospiClose], ['kospiClose'], 50, 10)
fin_func.figureSet(timeline, [kosdaqClose], ['kosdaqClose'], 50, 5)
fin_func.figureSet(timeline, [djiaClose], ['djiaClose'], 50, 100)
fin_func.figureSet(timeline, [spClose], ['spClose'], 50, 100)
fin_func.figureSet(timeline, [nasdaqClose], ['nasdaqClose'], 50, 100)
fin_func.figureSet(timeline, [kuClose], ['kuClose'], 50, 10)
fin_func.figureSet(timeline, [kospiClose, kosdaqClose, spClose, kuClose], ['kospiClose', 'kosdaqClose', 'spClose', 'kuClose'], 50, 10)
fin_func.figureSet(timeline, [djiaClose, nasdaqClose], ['djiaClose', 'nasdaqClose'], 50, 100)


